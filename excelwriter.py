import os
import sys
from openpyxl import load_workbook, Workbook

class ExcelFile:
    def __init__(self, filename):
        self.file_path = self.get_resource_path(filename)
        if os.path.exists(self.file_path):
            self.workbook = load_workbook(self.file_path)
        else:
            self.workbook = Workbook()
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])

    def get_resource_path(self, relative_path):
        if getattr(sys, 'frozen', False):
            exe_dir = os.path.dirname(sys.executable)
            file_path = os.path.join(exe_dir, relative_path)
            if os.path.exists(file_path):
                return file_path
            try:
                base_path = sys._MEIPASS
            except AttributeError:
                base_path = exe_dir
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_path, relative_path)
    
    def save(self):
        self.workbook.save(self.file_path)
    
    def get_sheet(self, sheet):
        if sheet in self.workbook.sheetnames:
            return self.workbook[sheet]
        return self.workbook.create_sheet(title=sheet)
    
    def clear_sheet(self, sheet):
        ws = self.get_sheet(sheet)
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    
    def apply_format(self, sheet="", position=None, format_code=None):
        """
        Apply Excel number format to cell(s).
        
        Args:
            sheet: Sheet name
            position: (row, col) tuple for single cell, or ((start_row, start_col), (end_row, end_col)) for range
            format_code: Excel format code (e.g., "#,##0", "yyyy-mm-dd", "General", "0.00")
        
        Examples:
            apply_format("Sheet1", (1, 1), "#,##0")  # Number with thousands separator
            apply_format("Sheet1", (1, 1), "yyyy-mm-dd")  # Date format
            apply_format("Sheet1", ((1, 1), (10, 5)), "0.00")  # Range format
        """
        ws = self.get_sheet(sheet)
        if format_code is None:
            return
        
        if position is None:
            return
        
        # Check if it's a range (tuple of tuples)
        if isinstance(position[0], tuple) and isinstance(position[1], tuple):
            start_row, start_col = position[0]
            end_row, end_col = position[1]
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws.cell(row, col).number_format = format_code
        else:
            # Single cell
            row, col = position
            ws.cell(row, col).number_format = format_code
    
    def write_rows_dict(self, sheet="", position=(1,1), datas=[], headers=[], show_headers=True, format=None):
        # datas : list of dicts, using headers as keys
        ws = self.get_sheet(sheet)
        start_row, start_col = position
        row = start_row
        if show_headers:
            for col, header in enumerate(headers, start_col):
                ws.cell(row, col).value = header
            row += 1
        for data in datas:
            for col, header in enumerate(headers, start_col):
                cell = ws.cell(row, col)
                cell.value = data.get(header)
                if format: cell.number_format = format
            row += 1
    
    def write_cols_dict(self, sheet="", position=(1, 1), datas={}, headers=[], show_headers=True, format=None):
        # datas : dict of lists, using headers as keys
        ws = self.get_sheet(sheet)
        start_row, start_col = position
        data_start_row = start_row + 1 if show_headers else start_row
        for col, header in enumerate(headers, start_col):
            if show_headers:
                ws.cell(start_row, col).value = header
            for row, value in enumerate(datas.get(header, []), data_start_row):
                cell = ws.cell(row, col)
                cell.value = value
                if format: cell.number_format = format
    
    def write_cell(self, sheet="", position=(1,1), content=None, format=None):
        ws = self.get_sheet(sheet)
        row, col = position
        cell = ws.cell(row, col)
        cell.value = content
        if format: cell.number_format = format
    
    def write_table_list(self, sheet="", position=(1,1), datas=[], format=None):
        # datas : list of lists
        ws = self.get_sheet(sheet)
        start_row, start_col = position
        for row_idx, data_row in enumerate(datas):
            for col_idx, value in enumerate(data_row):
                cell = ws.cell(start_row + row_idx, start_col + col_idx)
                cell.value = value
                if format: cell.number_format = format
